#!/usr/bin/python3.4
# -*- coding=utf-8 -*-
#本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
#教主QQ:605658506
#亁颐堂官网www.qytang.com
#乾颐盾是由亁颐堂现任明教教主开发的综合性安全课程
#包括传统网络安全（防火墙，IPS...）与Python语言和黑客渗透课程！

from openpyxl import Workbook
from openpyxl import load_workbook

dict_excel = {'test123':('cisco123',15), 'test456':('cisco456',1), 'test789':('cisco789',1)}

def excel_write(file = 'write_pyxl.xlsx', sheel_name = 'Sheet1', write_dict = dict_excel):
	wb = Workbook()
	ws = wb.active
	#ws = wb.create_sheet()
	#ws.title = "QYT PyXL"
	ws['A1'] = '用户'
	ws['B1'] = '密码'
	ws['C1'] = '级别'
	row_location = 2 
	for x,y in write_dict.items():
		user_locatin = 'A' + str(row_location)
		pass_locatin = 'B' + str(row_location)
		priv_locatin = 'C' + str(row_location)
		ws[user_locatin] = x
		ws[pass_locatin] = y[0]
		ws[priv_locatin] = y[1]
		row_location += 1 
	wb.save(file)

if __name__ == "__main__":	
	excel_write()
