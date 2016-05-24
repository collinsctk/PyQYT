#!/usr/bin/python3.4
# -*- coding=utf-8 -*-
#本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
#教主QQ:605658506
#亁颐堂官网www.qytang.com
#乾颐盾是由亁颐堂现任明教教主开发的综合性安全课程
#包括传统网络安全（防火墙，IPS...）与Python语言和黑客渗透课程！

from openpyxl import Workbook
from openpyxl import load_workbook

def excel_parser_return_dict(file = 'test.xlsx', sheel_name = 'Sheet1'):
	data = load_workbook(file)
	table = data[sheel_name]
	#print('%-22s %-22s %s' % (table['A1'].value, table['B1'].value, table['C1'].value))
	#print(table.rows)
	excel_dict = {}
	row_location = 0 
	for row in table.iter_rows():
		if row_location == 0:
			row_location += 1
			continue
		else:
			cell_location = 0
			for cell in row:
				if cell_location == 0:
					tmp_user = cell.value
					cell_location += 1
				elif cell_location == 1:
					tmp_pass = cell.value
					cell_location += 1
				elif cell_location == 2:
					tmp_priv = cell.value
					cell_location += 1
			excel_dict[tmp_user] = tmp_pass, tmp_priv
		row_location += 1
	return excel_dict

if __name__ == "__main__":
	print(excel_parser_return_dict('test.xlsx'))
