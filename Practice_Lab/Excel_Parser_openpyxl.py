#!/usr/bin/python3.4
# -*- coding=utf-8 -*-
#本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
#教主QQ:605658506
#亁颐堂官网www.qytang.com
#乾颐盾是由亁颐堂现任明教教主开发的综合性安全课程
#包括传统网络安全（防火墙，IPS...）与Python语言和黑客渗透课程！

from openpyxl import Workbook
from openpyxl import load_workbook

def excel_parser(file = 'test.xlsx', sheel_name = 'Sheet1'):
	data = load_workbook(file)
	table = data[sheel_name]
	print('%-22s %-22s %s' % (table['A1'].value, table['B1'].value, table['C1'].value))
	#print(table.rows)
	row_location = 0 
	for row in table.iter_rows():
		if row_location == 0:
			row_location += 1
			continue
		else:
			cell_location = 0
			for cell in row:
				if cell_location == 0:
					print('用户:%-20s' % cell.value, end='')
					cell_location += 1
				elif cell_location == 1:
					print('密码:%-20s' % cell.value, end='')
					cell_location += 1
				elif cell_location == 2:
					print('级别:%-20s' % cell.value, end='')
					cell_location += 1
			print()
		row_location += 1

if __name__ == "__main__":
	excel_parser('test.xlsx')
